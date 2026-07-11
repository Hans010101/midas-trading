"""P0 两个零 key 官方 JSON 源:Fed calendar.json + BEA release_dates.json(每日刷)。

★解析层纯函数(parse_*·喂 dict 可单测·不碰网络);拉取层 async httpx(照 email/oxapay 范式)。
★实测形状(2026-07-10 亲手 curl,调研 verified):
  - Fed:{"events":[2577 条],"announcement"} · ★文件带 BOM(utf-8-sig)· type=="FOMC" 135 条,
    title 有大小写/前导空格变体("FOMC Meeting"/"FOMC meeting"/" FOMC Minutes")须 normalize;
    会议条目 month="2026-12"+days="9"(两日会的【末日=决议日】,范围如 "17-18" 取末数);
    time="2:00 p.m."(★无时区标注=美东,America/New_York 换算,项目铁律 tz-aware)。
  - BEA:dict{release 名 → {"release_dates":[ISO8601 带时区]}} · 个别重复日期需去重。
🔴 红线:纯日程数据 · 零 key 零凭证 · 与交易逻辑零关联。
"""

from __future__ import annotations

import io
import logging
import re
import xml.etree.ElementTree as ET  # noqa: N817 · 标准库惯用别名
from collections.abc import Callable
from datetime import UTC, date, datetime, time
from typing import Any
from zoneinfo import ZoneInfo

import httpx

logger = logging.getLogger(__name__)

FED_CALENDAR_URL = "https://www.federalreserve.gov/json/calendar.json"
BEA_RELEASE_DATES_URL = "https://apps.bea.gov/API/signup/release_dates.json"
# KOSTAT(2026 改名 국가데이터처/MODS)官方年度 xlsx · 路径纯约定按年份填充(调研实测 200)
KOSTAT_XLSX_URL = "https://mods.go.kr/ansk/file/schedule_{year}.xlsx"
# 日本:総務省統計局 e-Stat 公表予定 XML(★UTF-16 编码 · 调研实测坑)+ BOJ 统计公表予定 xlsx
JP_ESTAT_CPI_URL = "https://www.stat.go.jp/data/kouhyou/e-stat_cpi.xml"
JP_ESTAT_ROUDOU_URL = "https://www.stat.go.jp/data/kouhyou/e-stat_roudou.xml"
BOJ_SCHEDULE_XLSX_URL = "https://www.boj.or.jp/statistics/outline/tkohyos.xlsx"
# 零 key 公共源 · UA 自我标识(公共数据礼仪·非绕反爬)
_UA = {"User-Agent": "MidasTerminal/1.0 (data-schedule; contact via midastrade.asia)"}
_TIMEOUT = 30.0

_ET = ZoneInfo("America/New_York")
_KST = ZoneInfo("Asia/Seoul")
_JST = ZoneInfo("Asia/Tokyo")
_LONDON = ZoneInfo("Europe/London")
_CET = ZoneInfo("Europe/Berlin")   # 德/法/意同 CET 偏移(仅用于日期不跨 CST 午夜,10:00 本地存储)

# BEA release 名 → (event_type, 中文标题, importance)· ★最小集:GDP + PCE(FOMC 关注的通胀口径)
BEA_RELEASES: dict[str, tuple[str, str, int]] = {
    "Gross Domestic Product": ("us_gdp", "美国GDP", 2),
    "Personal Income and Outlays": ("us_pce", "美国PCE物价", 2),
}

# KOSTAT 年表 보도자료명(标题)含此韩文子串 → (event_type, 中文标题)· ★最小集:三大月度指标
#   소비자물가동향=CPI · 고용동향=就业 · 산업활동동향=产业活动(2026 xlsx 实测各 12 条,08:00 KST)
KOSTAT_INDICATORS: tuple[tuple[str, str, str], ...] = (
    ("소비자물가동향", "kr_cpi", "韩国CPI"),
    ("고용동향", "kr_employment", "韩国就业动向"),
    ("산업활동동향", "kr_ind_activity", "韩国产业活动"),
)


def _parse_fed_time(raw: str) -> tuple[int, int]:
    """'2:00 p.m.' → (14, 0) · 解析不了回退 (14, 0)(FOMC 决议惯例时刻)。"""
    m = re.match(r"(\d{1,2})(?::(\d{2}))?\s*([ap])\.?m", raw.strip().lower())
    if not m:
        return (14, 0)
    hour = int(m.group(1)) % 12
    if m.group(3) == "p":
        hour += 12
    return (hour, int(m.group(2) or 0))


def parse_fed_events(data: dict[str, Any]) -> list[dict[str, Any]]:
    """Fed calendar.json → FOMC 利率决议事件列表(纯函数 · 可单测)。

    ★只取 title normalize 后 == "fomc meeting" 的条目(决议 ★3);Minutes/Press Conference/
    Speeches 不进最小集(别过度扩张)。days 可为范围("17-18")→ 取末数 = 两日会末日 = 决议日。
    """
    out: list[dict[str, Any]] = []
    for ev in data.get("events", []):
        title = str(ev.get("title", "")).strip().lower()
        if ev.get("type") != "FOMC" or title != "fomc meeting":
            continue
        month = str(ev.get("month", ""))          # "2026-12"
        days = str(ev.get("days", "")).strip()    # "9" 或 "17-18"
        last_day = days.split("-")[-1].strip()
        if not re.fullmatch(r"\d{4}-\d{2}", month) or not last_day.isdigit():
            continue
        hour, minute = _parse_fed_time(str(ev.get("time", "")))
        try:
            local = datetime(
                int(month[:4]), int(month[5:7]), int(last_day), hour, minute, tzinfo=_ET,
            )
        except ValueError:
            continue
        scheduled = local.astimezone(UTC)
        out.append({
            "event_key": f"fomc-{local:%Y-%m-%d}",
            "event_type": "fomc",
            "title": "FOMC 利率决议",
            "markets": ["us", "crypto", "hk", "cn"],
            "importance": 3,
            "scheduled_at": scheduled,
            "time_confirmed": True,
            "source": "fed_json",
        })
    return out


def parse_bea_events(data: dict[str, Any]) -> list[dict[str, Any]]:
    """BEA release_dates.json → GDP/PCE 事件列表(纯函数 · ISO 带时区 · 去重)。"""
    out: list[dict[str, Any]] = []
    seen: set[str] = set()
    for release_name, (etype, title, importance) in BEA_RELEASES.items():
        entry = data.get(release_name) or {}
        for iso in entry.get("release_dates", []):
            try:
                dt = datetime.fromisoformat(str(iso))
            except ValueError:
                continue
            if dt.tzinfo is None:  # 铁律:绝不留 naive
                dt = dt.replace(tzinfo=UTC)
            key = f"{etype}-{dt.astimezone(UTC):%Y-%m-%d}"
            if key in seen:  # ★个别重复日期(实测)去重
                continue
            seen.add(key)
            out.append({
                "event_key": key,
                "event_type": etype,
                "title": title,
                "markets": ["us", "crypto"] if etype == "us_pce" else ["us"],
                "importance": importance,
                "scheduled_at": dt.astimezone(UTC),
                "time_confirmed": True,
                "source": "bea_json",
            })
    return out


async def fetch_fed_events() -> list[dict[str, Any]]:
    """拉 Fed calendar.json(★BOM → utf-8-sig)→ 解析 FOMC。失败抛给调用方记 fail。"""
    async with httpx.AsyncClient(timeout=_TIMEOUT, headers=_UA) as client:
        resp = await client.get(FED_CALENDAR_URL)
        resp.raise_for_status()
        import json  # noqa: PLC0415 · 手动 utf-8-sig 解码(resp.json() 会被 BOM 噎住)

        data = json.loads(resp.content.decode("utf-8-sig"))
    events = parse_fed_events(data)
    logger.info("[econ-cal] fed_json 解析 FOMC 决议 %d 条", len(events))
    return events


async def fetch_bea_events() -> list[dict[str, Any]]:
    """拉 BEA release_dates.json → 解析 GDP/PCE。失败抛给调用方记 fail。"""
    async with httpx.AsyncClient(timeout=_TIMEOUT, headers=_UA) as client:
        resp = await client.get(BEA_RELEASE_DATES_URL)
        resp.raise_for_status()
        import json  # noqa: PLC0415

        data = json.loads(resp.content.decode("utf-8-sig"))
    events = parse_bea_events(data)
    logger.info("[econ-cal] bea_json 解析 %d 条", len(events))
    return events


# ── KOSTAT(韩国国家数据处/MODS)年度 xlsx ─────────────────────────────────────
# ★实测形状(2026-07-10 亲手下载解包):单 sheet · 第 3 行表头(보도일자/보도시간/
#   보도자료명/담당과)· 数据从第 4 行 · 日期 "M.D.(요일)" · 时刻 "08:00" 或 datetime.time ·
#   三大月度指标各 12 条 08:00 KST。🔴 韩国全 importance=1 + markets=["kr"](红线不注入决策卡)。


def _parse_kostat_time(raw: Any) -> tuple[int, int]:
    """'08:00' / datetime.time → (8, 0) · 解析不了回退 (8, 0)(KOSTAT 惯例发布时刻)。"""
    if isinstance(raw, time):
        return (raw.hour, raw.minute)
    m = re.match(r"(\d{1,2}):(\d{2})", str(raw).strip())
    return (int(m.group(1)), int(m.group(2))) if m else (8, 0)


def parse_kostat_rows(rows: list[tuple[Any, ...]], year: int) -> list[dict[str, Any]]:
    """KOSTAT 年表行(보도일자, 보도시간, 보도자료명, …)→ 三大指标事件(纯函数 · 可单测)。

    只取标题含 KOSTAT_INDICATORS 韩文子串的行(别过度扩张:全表 199 条,只要 CPI/就业/
    产业活动 36 条)· 日期 "M.D.(요일)" 取 M.D + 传入 year · 韩文星期忽略(用 year+M+D)。
    """
    out: list[dict[str, Any]] = []
    seen: set[str] = set()
    for row in rows:
        if len(row) < 3 or not row[0] or not row[2]:
            continue
        date_raw, time_raw, title_raw = str(row[0]).strip(), row[1], str(row[2])
        # ★先匹配指标(表头/其他报告类型静默跳过是对的);再解析日期——这样「三大指标却
        #   日期解析失败」能留 warn(否则静默漏采:同一 xlsx 里日期格式按录入人不一致,
        #   实测有 "8.31(월)" 无尾点 / "9. 8.(화)" 带空格等 8 种变体,对抗自审实锤)。
        # ★结尾匹配(非子串):主报告标题恒为「…M월[및 연간]<指标>동향」,指标名在结尾。
        #   子串匹配会把未来年「청년층 고용동향 부가조사」等含名副报告误采(交叉审 P2);
        #   endswith 收紧 → 真 xlsx 36 条零变化,副报告排除(实测)。
        stripped_title = title_raw.rstrip()
        etype = title = None
        for needle, et, zh in KOSTAT_INDICATORS:
            if stripped_title.endswith(needle):
                etype, title = et, zh
                break
        if etype is None:
            continue
        # 容忍尾点缺失("8.31(월)")与内嵌空格("9. 8.(화)")· "1.14.(수)" 仍匹配
        dm = re.match(r"(\d{1,2})\.\s*(\d{1,2})\.?", date_raw)
        if not dm:
            logger.warning("[econ-cal] kostat 指标行日期无法解析(疑格式漂移·漏采): %r", date_raw)
            continue
        month, day = int(dm.group(1)), int(dm.group(2))
        hour, minute = _parse_kostat_time(time_raw)
        try:
            local = datetime(year, month, day, hour, minute, tzinfo=_KST)
        except ValueError:
            continue
        key = f"{etype}-{year}-{month:02d}-{day:02d}"
        if key in seen:   # 同指标同日去重(年表偶有重复行)
            continue
        seen.add(key)
        out.append({
            "event_key": key,
            "event_type": etype,
            "title": title,
            "markets": ["kr"],      # ★非四市场 → 决策卡永不注入(叠加 importance=1)
            "importance": 1,        # ★韩国恒 1(红线:绝不提 2)
            "scheduled_at": local.astimezone(UTC),
            "time_confirmed": True,
            "source": "kostat",
        })
    return out


def _extract_kostat_rows(content: bytes) -> list[tuple[Any, ...]]:
    """xlsx 字节 → 首个 sheet 全部数据行(轻量:read_only)· openpyxl 已是显式依赖。"""
    import openpyxl  # noqa: PLC0415 · 仅此路径用,惰性 import 避免拖累无关调用

    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        return [tuple(r) for r in ws.iter_rows(values_only=True)]
    finally:
        wb.close()


async def fetch_kostat_events(year: int | None = None) -> list[dict[str, Any]]:
    """拉 KOSTAT schedule_{year}.xlsx → 解析三大指标。当年为主(失败抛出记 fail);
    临近年末顺带best-effort拉次年(404 正常,静默忽略)· 失效模式良性:存量 30 天仍有效。
    """
    now_kst = datetime.now(tz=_KST)
    base_year = year or now_kst.year
    events = await _fetch_kostat_year(base_year)          # 主年:失败向上抛
    if year is None and now_kst.month >= 10:              # 10 月起次年表可能已出,顺带拉
        try:
            events += await _fetch_kostat_year(base_year + 1)
        except Exception as exc:  # noqa: BLE001 · 次年 404/未发布正常,不影响主年
            logger.info("[econ-cal] kostat 次年 %d 暂不可用(正常): %s", base_year + 1, exc)
    logger.info("[econ-cal] kostat 解析 %d 条", len(events))
    return events


async def _fetch_kostat_year(year: int) -> list[dict[str, Any]]:
    async with httpx.AsyncClient(timeout=_TIMEOUT, headers=_UA) as client:
        resp = await client.get(KOSTAT_XLSX_URL.format(year=year))
        resp.raise_for_status()
    return parse_kostat_rows(_extract_kostat_rows(resp.content), year)


# ── 日本 · 総務省統計局 e-Stat 公表予定 XML(CPI / 失業率)──────────────────────
# ★实测形状(2026-07-11 亲手下载解包):★★UTF-16 LE 编码(BOM · resp.text 会误按 UTF-8
#   噎成乱码 → 必须 content.decode("utf-16"))· os_code>class_1>class_2(数据期)>…>class_5
#   叶承载 release_year/month/day/hour/minute(精确到分)+ internet_url。滚动约 9 个月未来。
#   同一 os_code 内混有非目标发布(CPI 東京都区部速報 / 遡及結果、失業率 詳細集計 14:00)→
#   keep 谓词按 class_1/class_2 名筛出目标(全国月度 CPI / 基本集計 失業率)。
# 🔴 日本全 importance=1 + markets=["jp"](非四交易市场 → 决策卡永不注入,同韩国双重锁)。


def _safe_int(raw: str | None, default: int) -> int:
    """不可信文本 → int · 非数字(「未定」/空白/None)回退 default(绝不裸抛拖垮整份调查)。"""
    try:
        return int(raw) if raw and raw.strip() else default
    except (TypeError, ValueError):
        return default


def _cpi_keep(class_1_name: str, class_2_name: str) -> bool:
    """全国(非東京都区部速報)· 月度数据期(排除「2025年基準…遡及…接続指数」等特殊发布)。"""
    return class_1_name == "全国" and class_2_name.endswith("月分")


def _unemp_keep(class_1_name: str, class_2_name: str) -> bool:  # noqa: ARG001
    """基本集計(失業率头条 · 08:30)· 排除詳細集計(季度 14:00 · 非市场事件)。"""
    return class_2_name.startswith("基本集計")


# (os_code name 校验, event_type, 中文标题, keep 谓词)· os_code 校验防路径漂移/换文件
JP_ESTAT_SPECS: tuple[tuple[str, str, str, Callable[[str, str], bool]], ...] = (
    ("消費者物価指数", "jp_cpi", "日本CPI", _cpi_keep),
    ("労働力調査", "jp_unemp", "日本失业率", _unemp_keep),
)


def parse_estat_xml(
    content: bytes, *, os_name: str, event_type: str, title: str,
    keep: Callable[[str, str], bool],
) -> list[dict[str, Any]]:
    """統計局 e-Stat XML(UTF-16 bytes)→ 目标指标未来发布(纯函数 · 同日去重 · 可单测)。"""
    text = content.decode("utf-16")                       # ★UTF-16(别当 UTF-8)
    text = re.sub(r"^\s*<\?xml[^>]*\?>", "", text, count=1)  # ET 拒带 encoding 声明的 str
    try:
        root = ET.fromstring(text)                        # noqa: S314 · 官方源·非不可信输入
    except ET.ParseError as exc:
        logger.warning("[econ-cal] estat %s XML 解析失败(疑源漂移): %s", event_type, exc)
        return []
    os_el = root.find("os_code")
    if os_el is None or os_el.get("name") != os_name:
        got = os_el.get("name") if os_el is not None else None
        logger.warning("[econ-cal] estat os_code 不符(疑源漂移):期望 %r 实得 %r", os_name, got)
        return []
    seen: dict[str, tuple[int, int]] = {}                 # date_key → (h,m)· 同日首见胜(头条 08:30)
    for c1 in root.iter("class_1"):
        c1n = c1.get("name") or ""
        for c2 in c1.iter("class_2"):
            c2n = c2.get("name") or ""
            if not keep(c1n, c2n):
                continue
            for c5 in c2.iter("class_5"):
                def _g(tag: str, el: ET.Element = c5) -> str | None:
                    found = el.find(tag)
                    return found.text if found is not None else None
                y, m, d = _g("release_year"), _g("release_month"), _g("release_day")
                if not (y and m and d):
                    continue
                try:
                    key = f"{int(y):04d}-{int(m):02d}-{int(d):02d}"
                except ValueError:
                    continue
                hh_raw, mm_raw = _g("release_hour"), _g("release_minute")
                # ★时刻字段与日期字段同属不可信输入,必须同样容错:非数字(如「未定」/
                #   空白)不能裸 int() 抛异常拖垮整份调查(对抗自审 P2 · 只丢该行的粒度)
                seen.setdefault(key, (_safe_int(hh_raw, 8), _safe_int(mm_raw, 30)))
    out: list[dict[str, Any]] = []
    for key, (hh, mm) in seen.items():
        yy, mo, dd = (int(x) for x in key.split("-"))
        try:
            local = datetime(yy, mo, dd, hh, mm, tzinfo=_JST)
        except ValueError:
            continue
        out.append({
            "event_key": f"{event_type}-{key}",
            "event_type": event_type,
            "title": title,
            "markets": ["jp"],       # ★非四市场 → 决策卡永不注入(叠加 importance=1)
            "importance": 1,         # ★日本恒 1(红线:绝不提 2)
            "scheduled_at": local.astimezone(UTC),
            "time_confirmed": True,
            "source": "jp_estat",
        })
    return out


async def fetch_jp_estat_events() -> list[dict[str, Any]]:
    """拉統計局 CPI + 失業率 XML → 解析。★两源恒有未来发布(滚动 9 个月)→ 任一调查解析
    0 条 = os_code/class_1 名漂移(CPI 换基年重命名节点是真实场景)或断供 → 直接抛,让
    worker 不 mark_success → 3 天保鲜转 stale。绝不「部分成功也 mark_success」掩盖单调查
    静默停更(对抗自审 P2:旧版一份 XML 漂移、另一份正常时,聚合非空→假新鲜→漂移无人知)。
    """
    out: list[dict[str, Any]] = []
    async with httpx.AsyncClient(timeout=_TIMEOUT, headers=_UA) as client:
        for url, (os_name, etype, title, keep) in zip(
            (JP_ESTAT_CPI_URL, JP_ESTAT_ROUDOU_URL), JP_ESTAT_SPECS, strict=True,
        ):
            resp = await client.get(url)                  # 网络/404 自然抛 → 记 fail
            resp.raise_for_status()
            evs = parse_estat_xml(
                resp.content, os_name=os_name, event_type=etype, title=title, keep=keep,
            )
            if not evs:                                    # ★恒有未来发布 · 0 条=漂移/断供
                msg = f"jp_estat {etype} 解析 0 条(疑 os_code/class_1 漂移),不 mark_success"
                raise ValueError(msg)
            out += evs
    logger.info("[econ-cal] jp_estat 解析 %d 条", len(out))
    return out


# ── 日本 · 日本銀行 BOJ 統計公表予定 xlsx(短観 Tankan · 非议息)────────────────────
# ★实测形状(2026-07-11 亲手下载解包):2 sheet,第 2 sheet「統計データ」308 行 = 统计 ×
#   月份矩阵。每统计两行:时刻行(col3="08:50:00")+ 日期行(col3="(四半期)"·月列含原生
#   datetime `2026-10-01`,无需令和转换)· 同 col1 名配对。短観概要 08:50 JST = 市场事件。


# (col1 統計名子串, event_type, 中文标题)· 短観概要及び要旨(头条速報 · 非調査全容/時系列)
BOJ_STATS: tuple[tuple[str, str, str], ...] = (
    ("短観（全国企業短期経済観測調査）／概要", "jp_tankan", "日本短观Tankan"),
)


def parse_boj_rows(rows: list[tuple[Any, ...]]) -> list[dict[str, Any]]:
    """BOJ 統計データ sheet 全行 → 短観等目标统计未来发布(纯函数 · 同日去重 · 可单测)。

    每统计的时刻在「时刻行」col3、日期在「日期行」月列(原生 datetime)· 同 col1 名配对。
    """
    out: list[dict[str, Any]] = []
    for needle, etype, title in BOJ_STATS:
        time_hm: tuple[int, int] | None = None
        dates: list[date] = []
        for row in rows:
            name = str(row[1]) if len(row) > 1 and row[1] else ""
            if needle not in name:
                continue
            c3 = str(row[3]) if len(row) > 3 and row[3] else ""
            tm = re.match(r"(\d{1,2}):(\d{2})", c3)       # 时刻行 "08:50:00"
            if tm:
                time_hm = (int(tm.group(1)), int(tm.group(2)))
            for cell in row[4:]:                          # 月列:原生 datetime = 排定发布
                if isinstance(cell, datetime):
                    dates.append(cell.date())
                elif isinstance(cell, date):
                    dates.append(cell)
        hh, mm = time_hm or (8, 50)                        # 短観概要惯例 08:50 JST
        seen: set[str] = set()
        for dd in dates:
            key = f"{dd:%Y-%m-%d}"
            if key in seen:
                continue
            seen.add(key)
            local = datetime(dd.year, dd.month, dd.day, hh, mm, tzinfo=_JST)
            out.append({
                "event_key": f"{etype}-{key}",
                "event_type": etype,
                "title": title,
                "markets": ["jp"],   # ★非四市场 → 决策卡永不注入(叠加 importance=1)
                "importance": 1,     # ★日本恒 1(红线:绝不提 2)
                "scheduled_at": local.astimezone(UTC),
                "time_confirmed": True,
                "source": "boj_xlsx",
            })
    return out


def _extract_boj_rows(content: bytes) -> list[tuple[Any, ...]]:
    """xlsx 字节 →「統計データ」sheet 全行(无则退回末个 sheet)· openpyxl 惰性 import。"""
    import openpyxl  # noqa: PLC0415

    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        name = "統計データ" if "統計データ" in wb.sheetnames else wb.sheetnames[-1]
        return [tuple(r) for r in wb[name].iter_rows(values_only=True)]
    finally:
        wb.close()


async def fetch_boj_events() -> list[dict[str, Any]]:
    """拉 BOJ tkohyos.xlsx → 解析短観。失败抛给调用方记 fail(存量 30 天仍有效)。"""
    async with httpx.AsyncClient(timeout=_TIMEOUT, headers=_UA) as client:
        resp = await client.get(BOJ_SCHEDULE_XLSX_URL)
        resp.raise_for_status()
    events = parse_boj_rows(_extract_boj_rows(resp.content))
    logger.info("[econ-cal] boj_xlsx 解析短観 %d 条", len(events))
    return events


# ── 欧洲四国(英/德/法/意)· IMF DSBB SDDS Advance Release Calendar ─────────────────
# ★实测形状(2026-07-11 亲手 curl · getARCReportList JSON):数组,每记录 = 国×指标,
#   MonthValues = 定长 13 槽数组,★槽下标 i = 发布【日历月】(1..12),非数据期;
#   槽值 {"Day":"22","Period":"Jun/26"} · Period = 数据参考期(仅信息,解析日期【不用】);
#   ★★发布日 = (槽下标月, 年) + Day —— 已用 ONS(英)+ ISTAT(意)权威日程逐日对表验证
#   (研究文档的 Period+Day 解码【错位一个月】,ONS 实证纠正)。年份用 getAdvanceMonths
#   有序列表跨年 wrap 推导。Day 三形态:普通"22" / "NLT 25"(不晚于·日不确定) /
#   "1,30"(意大利同月双发布·逗号拆·均落槽月)。DSBB 只到日无时刻 → 存本地 10:00 +
#   time_confirmed=False(不编造时刻·显示「时刻待定」)。
# 🔴 四国全 importance=1 + markets=["eu"](非四交易市场 → 决策卡永不注入,同韩日双重锁)。

DSBB_BASE = "https://dsbb.imf.org/api/customquery"
# ★未文档化端点走 Akamai WAF:必须浏览器 UA(自标识 UA 会被挡)
_DSBB_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/120.0 Safari/537.36"
    ),
    "Accept": "application/json",
}
# (ISO3, 国别前缀, 中文国名, tz)· 英国按 Hans 定归欧洲桶
DSBB_COUNTRIES: tuple[tuple[str, str, str, ZoneInfo], ...] = (
    ("GBR", "gb", "英国", _LONDON),
    ("DEU", "de", "德国", _CET),
    ("FRA", "fr", "法国", _CET),
    ("ITA", "it", "意大利", _CET),
)
# (DSBB Category code, 指标后缀, 中文指标名)· CPI 月度恒有未来条(保鲜锚)· GDP/部分失业率季度
DSBB_CATEGORIES: tuple[tuple[str, str, str], ...] = (
    ("CPI00", "cpi", "CPI"),
    ("NAG00", "gdp", "GDP"),
    ("UEM00", "unemp", "失业率"),
)


def build_dsbb_year_map(advance_months: list[dict[str, Any]], now: datetime) -> dict[int, int]:
    """getAdvanceMonths 有序 [{MonthNum,MonthText}] → {月:年}· 跨年 wrap(月号回落即 +1 年)。"""
    year = now.year
    prev: int | None = None
    out: dict[int, int] = {}
    for item in advance_months:
        try:
            mn = int(item["MonthNum"])
        except (KeyError, TypeError, ValueError):
            continue
        if prev is not None and mn < prev:
            year += 1
        out[mn] = year
        prev = mn
    return out


def _parse_dsbb_days(raw: str) -> list[tuple[int, bool]]:
    """Day 字段 → [(日, time_confirmed)]· "22"→普通· "NLT 25"→日不确定· "1,30"→双发布同槽月。

    ★DSBB 无时刻 → time_confirmed 恒 False(不编造);NLT 日本身也不确定,同 False。
    """
    out: list[tuple[int, bool]] = []
    for tok_raw in raw.split(","):
        tok = tok_raw.strip()
        if tok.upper().startswith("NLT"):
            tok = tok[3:].strip()      # "NLT 25" → "25"(不晚于日)
        if tok.isdigit():
            out.append((int(tok), False))  # DSBB 全 time_confirmed=False
    return out


def parse_dsbb_records(
    records: list[dict[str, Any]], year_map: dict[int, int],
) -> list[dict[str, Any]]:
    """DSBB getARCReportList 数组 → 事件(纯函数 · 槽下标=发布月 · 同日去重 · 可单测)。"""
    cty = {iso: (pre, zh, tz) for iso, pre, zh, tz in DSBB_COUNTRIES}
    cat = {code: (suf, zh) for code, suf, zh in DSBB_CATEGORIES}
    out: list[dict[str, Any]] = []
    seen: set[str] = set()
    for rec in records:
        if not isinstance(rec, dict):
            continue
        c, k = rec.get("CountryCode"), rec.get("CategoryCode")
        if c not in cty or k not in cat:
            continue
        prefix, country_zh, tz = cty[c]
        suffix, cat_zh = cat[k]
        etype = f"{prefix}_{suffix}"
        title = f"{country_zh}{cat_zh}"
        for i, slot in enumerate(rec.get("MonthValues") or []):
            if not slot or i not in year_map:   # 槽下标 = 发布日历月;不在活跃窗口跳过
                continue
            for day, conf in _parse_dsbb_days(str(slot.get("Day", ""))):
                try:
                    # 10:00 本地(无时刻·存以保 CST 显示日期不跨午夜)
                    local = datetime(year_map[i], i, day, 10, 0, tzinfo=tz)
                except ValueError:
                    continue
                key = f"{etype}-{year_map[i]}-{i:02d}-{day:02d}"
                if key in seen:
                    continue
                seen.add(key)
                out.append({
                    "event_key": key,
                    "event_type": etype,
                    "title": title,
                    "markets": ["eu"],   # ★非四市场 → 决策卡永不注入(叠加 importance=1)
                    "importance": 1,     # ★欧洲恒 1(红线:绝不提 2)
                    "scheduled_at": local.astimezone(UTC),
                    "time_confirmed": conf,   # DSBB 无时刻 → False(显示「时刻待定」)
                    "source": "dsbb",
                })
    return out


async def fetch_dsbb_events(now: datetime | None = None) -> list[dict[str, Any]]:
    """拉 DSBB 四国×三指标(12 次小 GET)+ getAdvanceMonths 定年。CPI 月度恒有未来条 →
    0 条 = WAF/端点漂移 → 抛(worker 不 mark_success 转 stale)· 同 jp_estat 范式。
    """
    now = now or datetime.now(tz=UTC)
    out: list[dict[str, Any]] = []
    async with httpx.AsyncClient(timeout=_TIMEOUT, headers=_DSBB_HEADERS) as client:
        am_resp = await client.get(
            f"{DSBB_BASE}/getAdvanceMonths",
            params={"Countries": "GBR", "Categories": "CPI00"},
        )
        am_resp.raise_for_status()
        year_map = build_dsbb_year_map(am_resp.json(), now)
        for iso, _pre, _zh, _tz in DSBB_COUNTRIES:
            for code, _suf, _czh in DSBB_CATEGORIES:
                resp = await client.get(
                    f"{DSBB_BASE}/getARCReportList",
                    params={"Countries": iso, "Categories": code},
                )
                resp.raise_for_status()
                data = resp.json()
                if isinstance(data, list):
                    out += parse_dsbb_records(data, year_map)
    if not out:                # 四国 CPI 恒有未来 · 0 条 = 端点/WAF 漂移
        msg = "dsbb 解析 0 条(疑 WAF/端点漂移),不 mark_success"
        raise ValueError(msg)
    logger.info("[econ-cal] dsbb 解析 %d 条", len(out))
    return out
